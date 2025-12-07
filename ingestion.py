import os
import json
import uuid
import re
from pathlib import Path
import fitz
import pdfplumber
from PIL import Image
import pytesseract
from docx import Document
import markdown
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from config import *
import io


USE_IMAGE_CAPTIONING = False  

if USE_IMAGE_CAPTIONING:
    from transformers import BlipProcessor, BlipForConditionalGeneration
    import torch
    _caption_processor = None
    _caption_model = None

    def get_caption_model():
        global _caption_processor, _caption_model
        if _caption_processor is None:
            _caption_processor = BlipProcessor.from_pretrained("Salesforce/blip-image-captioning-base") 
            _caption_model = BlipForConditionalGeneration.from_pretrained("Salesforce/blip-image-captioning-base").to("cpu")
        return _caption_processor, _caption_model


def table_to_natural_language(df, table_name):
    """
    Universal converter: Table → Natural Language
    Works for ANY table structure (sales, inventory, rankings, anything)
    """
    
    # Clean dataframe
    df = df.copy()
    df = df.dropna(how='all')
    
    # Clean column names and cell values
    df.columns = df.columns.astype(str).str.strip()
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.strip()
    
    # Start building natural language representation
    text = f"# Table: {table_name}\n\n"
    text += f"This table contains {len(df)} rows and {len(df.columns)} columns.\n\n"
    
    # Column Information ===
    text += "## Columns in this table:\n"
    for i, col in enumerate(df.columns, 1):
        col_type = "numeric" if pd.api.types.is_numeric_dtype(df[col]) else "text"
        text += f"{i}. {col} ({col_type})\n"
    text += "\n"
    
    # Row-by-Row Description ===
    text += "## Data entries (described row by row):\n\n"
    
    # Identify the primary column
    primary_col = df.columns[0]
    
    for idx, row in df.iterrows():

        primary_value = row[primary_col]
        text += f"**{primary_value}**:\n"
        
        # Describe each attribute
        for col in df.columns[1:]:
            value = row[col]
            if pd.notna(value) and str(value).strip() and str(value).lower() != 'nan':
                text += f"  - {col}: {value}\n"
        
        text += "\n"
    
    # Numeric Analysis ===
    numeric_cols = df.select_dtypes(include=['float64', 'int64']).columns.tolist()
    
    if numeric_cols:
        text += "## Numeric Insights:\n\n"
        
        for col in numeric_cols:
            if df[col].notna().any():
                text += f"### {col}:\n"
                
                text += f"- Lowest value: {df[col].min()}"
                min_entity = df.loc[df[col].idxmin(), primary_col]
                text += f" (for {min_entity})\n"
                
                text += f"- Highest value: {df[col].max()}"
                max_entity = df.loc[df[col].idxmax(), primary_col]
                text += f" (for {max_entity})\n"
                
                text += f"- Average: {df[col].mean():.2f}\n"
                text += f"- Total: {df[col].sum():.2f}\n"
                
                # Sort by this column 
                df_sorted_asc = df.sort_values(by=col).dropna(subset=[col])
                df_sorted_desc = df.sort_values(by=col, ascending=False).dropna(subset=[col])
                
                text += f"- Sorted from lowest to highest:\n"
                for rank, (_, r) in enumerate(df_sorted_asc.head(5).iterrows(), 1):
                    text += f"  {rank}. {r[primary_col]}: {r[col]}\n"
                
                text += f"- Sorted from highest to lowest:\n"
                for rank, (_, r) in enumerate(df_sorted_desc.head(5).iterrows(), 1):
                    text += f"  {rank}. {r[primary_col]}: {r[col]}\n"
                
                text += "\n"
    
    # Text Analysis
    text_cols = df.select_dtypes(include=['object']).columns.tolist()

    text_cols = [col for col in text_cols if col != primary_col]
    
    if text_cols:
        text += "## Text Attributes:\n\n"
        
        for col in text_cols:
            unique_values = df[col].dropna().unique()
            if len(unique_values) <= 20: 
                text += f"### {col}:\n"
                text += f"Unique values: {', '.join(map(str, unique_values))}\n\n"
                
                # Group by this column
                for val in unique_values:
                    matching = df[df[col] == val]
                    if len(matching) > 0:
                        text += f"- Entries with {col} = '{val}':\n"
                        for _, r in matching.head(5).iterrows():
                            text += f"  • {r[primary_col]}\n"
                        if len(matching) > 5:
                            text += f"  ... and {len(matching) - 5} more\n"
                        text += "\n"
    
    #  Relationship Patterns 
    if len(numeric_cols) >= 2:
        text += "## Relationships between columns:\n\n"
        
        # Find correlations
        try:
            corr_matrix = df[numeric_cols].corr()
            
            for i, col1 in enumerate(numeric_cols):
                for col2 in numeric_cols[i+1:]:
                    corr_value = corr_matrix.loc[col1, col2]
                    if abs(corr_value) > 0.5: 
                        if corr_value > 0:
                            text += f"- **{col1}** and **{col2}** are positively related (correlation: {corr_value:.2f})\n"
                            text += f"  When {col1} increases, {col2} also tends to increase.\n"
                        else:
                            text += f"- **{col1}** and **{col2}** are negatively related (correlation: {corr_value:.2f})\n"
                            text += f"  When {col1} increases, {col2} tends to decrease.\n"
            text += "\n"
        except:
            pass 
    
    #  Question-Answering Hints 
    text += "## Common queries this table can answer:\n\n"
    
    # Generic questions
    text += f"- How many {primary_col} are in this table? Answer: {len(df)}\n"
    text += f"- What are all the {primary_col}? Answer: {', '.join(df[primary_col].astype(str).tolist())}\n"
    
    # Numeric questions
    for col in numeric_cols:
        text += f"- What is the highest {col}? Answer: {df[col].max()} (for {df.loc[df[col].idxmax(), primary_col]})\n"
        text += f"- What is the lowest {col}? Answer: {df[col].min()} (for {df.loc[df[col].idxmin(), primary_col]})\n"
        text += f"- What is the average {col}? Answer: {df[col].mean():.2f}\n"
    
    text += "\n"
    
    # Raw Table 
    text += "## Raw table data:\n\n"
    text += "```\n"
    text += df.to_string(index=False, max_rows=50)
    if len(df) > 50:
        text += f"\n... and {len(df) - 50} more rows"
    text += "\n```\n"
    
    return text

def generate_caption(image_path):
    """Generate caption for an image - OPTIONAL for speed."""
    if not USE_IMAGE_CAPTIONING:
        return "Image caption disabled for faster processing"
    
    try:
        processor, model = get_caption_model()
        raw_image = Image.open(image_path).convert('RGB')
        raw_image.thumbnail((512, 512))
        inputs = processor(raw_image, return_tensors="pt").to(model.device)
        out = model.generate(**inputs, max_new_tokens=30) 
        caption = processor.decode(out[0], skip_special_tokens=True)
        return caption
    except Exception as e:
        print(f"⚠️ Captioning failed: {e}")
        return "Caption generation failed"

def ensure_doc_dirs(doc_id):
    """Create directory structure for a document."""
    base = os.path.join(BASE_DIR, doc_id)
    texts = os.path.join(base, TEXTS_SUBDIR)
    tables = os.path.join(base, TABLES_SUBDIR)
    images = os.path.join(base, IMAGES_SUBDIR)
    for p in (base, texts, tables, images):
        os.makedirs(p, exist_ok=True)
    return {"base": base, "texts": texts, "tables": tables, "images": images}

def chunk_text_words(text, chunk_size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    """Split text into overlapping word chunks."""
    if not text:
        return []
    words = re.split(r"\s+", text)
    chunks = []
    start = 0
    while start < len(words):
        end = min(start + chunk_size, len(words))
        chunk_text = " ".join(words[start:end]).strip()
        if chunk_text:
            chunks.append(chunk_text)
        if end == len(words):
            break
        start += chunk_size - overlap
    return chunks

#  Faster PDF extraction 
def extract_pdf(file_path, doc_dirs):
    """Extract text, tables, and images from PDF - OPTIMIZED."""
    text_joined = []
    tables_list = []
    image_paths = []


    try:
        with fitz.open(file_path) as pdf:
            for i, page in enumerate(pdf):
                page_num = i + 1
                text = page.get_text("text")
                
            
                if len(text.strip()) < 5:
                    pix = page.get_pixmap(dpi=150) 
                    img = Image.open(io.BytesIO(pix.tobytes()))
                    text = pytesseract.image_to_string(img)
                
                if text.strip():
                    text_joined.append({"page": page_num, "text": text})

                """
                for img_index, img in enumerate(page.get_images(full=True)):
                    xref = img[0]
                    base_image = pdf.extract_image(xref)
                    img_bytes = base_image["image"]
                    ext = base_image["ext"]
                    name = f"page_{page_num}_img_{img_index+1}.{ext}"
                    path = os.path.join(doc_dirs["images"], name)
                    with open(path, "wb") as f:
                        f.write(img_bytes)
                    image_paths.append({"page": page_num, "path": path})
                """
    except Exception as e:
        print(f"PDF text extraction error: {e}")

    # Use fitz for tables (faster than pdfplumber) 
    try:
        with fitz.open(file_path) as pdf:
            for i, page in enumerate(pdf):
                page_num = i + 1
                tables = page.find_tables()
                for j, table in enumerate(tables):
                    if table and len(table.extract()) > 1:
                        df = pd.DataFrame(table.extract()[1:], columns=table.extract()[0])
                        csv_name = f"table_page{page_num}_{j+1}.csv"
                        csv_path = os.path.join(doc_dirs["tables"], csv_name)
                        df.to_csv(csv_path, index=False)
                        tables_list.append({"page": page_num, "csv_path": csv_path, "df": df})
    except Exception as e:
        print(f"PDF table extraction error: {e}")

    return {"texts": text_joined, "tables": tables_list, "images": image_paths}

def extract_docx(file_path, doc_dirs):
    """Extract text and images from DOCX."""
    text_list = []
    images = []
    try:
        doc = Document(file_path)
        for p in doc.paragraphs:
            if p.text.strip():
                text_list.append({"page": 1, "text": p.text})
        

        """
        for i, rel in enumerate(doc.part.rels.values()):
            if "image" in rel.reltype:
                img_blob = rel.target_part.blob
                path = os.path.join(doc_dirs["images"], f"docx_img_{i+1}.png")
                with open(path, "wb") as f:
                    f.write(img_blob)
                images.append({"page": 1, "path": path})
        """
    except Exception as e:
        print(f"DOCX extraction error: {e}")
    return {"texts": text_list, "tables": [], "images": images}

def extract_txt(file_path, doc_dirs):
    """Extract text from TXT/MD files."""
    with open(file_path, "r", encoding="utf-8", errors='ignore') as f:
        raw = f.read()
    return {"texts": [{"page": 1, "text": raw}], "tables": [], "images": []}

def extract_csv(file_path, doc_dirs):
    """Extract CSV with universal smart text representation."""
    try:
        df = pd.read_csv(file_path)
        

        df.columns = df.columns.str.strip()
        
        csv_name = os.path.basename(file_path)
        csv_path = os.path.join(doc_dirs["tables"], csv_name)
        df.to_csv(csv_path, index=False)
        
        # Convert to natural language
        table_text = table_to_natural_language(df, csv_name)
        
        return {
            "texts": [{
                "page": 1,
                "text": table_text
            }],
            "tables": [{
                "sheet": "main",
                "csv_path": csv_path,
                "df": df
            }],
            "images": []
        }
    except Exception as e:
        print(f"CSV extraction error: {e}")
        return {"texts": [], "tables": [], "images": []}
    

def extract_xlsx(file_path, doc_dirs):
    """Extract Excel with universal smart text representation."""
    tables = []
    texts = []
    
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            data = [[cell.value for cell in row] for row in ws.iter_rows()]
            if not data or len(data) < 2:
                continue
            
            df = pd.DataFrame(data[1:], columns=data[0])
            df.columns = df.columns.str.strip()
            
            csv_name = f"{sheet}.csv"
            csv_path = os.path.join(doc_dirs["tables"], csv_name)
            df.to_csv(csv_path, index=False)
            
            tables.append({
                "sheet": sheet,
                "csv_path": csv_path,
                "df": df
            })
            
            # Convert to natural language
            table_text = table_to_natural_language(df, f"{sheet} ({csv_name})")
            texts.append({
                "page": sheet,
                "text": table_text
            })
            
    except Exception as e:
        print(f"XLSX extraction error: {e}")
    
    return {"texts": texts, "tables": tables, "images": []}


def ingest_file(file_bytes, filename, username=None):
    """Main ingestion function - OPTIMIZED with username tracking."""
    doc_id = str(uuid.uuid4())
    doc_dirs = ensure_doc_dirs(doc_id)
    

    temp_path = os.path.join(doc_dirs["base"], filename)
    with open(temp_path, "wb") as f:
        f.write(file_bytes)
    

    ext = os.path.splitext(filename)[1].lower()
    if ext == ".pdf":
        extracted = extract_pdf(temp_path, doc_dirs)
    elif ext == ".docx":
        extracted = extract_docx(temp_path, doc_dirs)
    elif ext in [".txt", ".md"]:
        extracted = extract_txt(temp_path, doc_dirs)
    elif ext == ".csv":
        extracted = extract_csv(temp_path, doc_dirs)
    elif ext in [".xlsx", ".xls"]:
        extracted = extract_xlsx(temp_path, doc_dirs)
    else:
        raise ValueError(f"Unsupported file type: {ext}")
    
    # Build metadata
    doc_meta = {
        "doc_id": doc_id,
        "filename": filename,
        "ext": ext,
        "base_dir": doc_dirs["base"],
        "text_files_count": len(extracted.get("texts", [])),
        "tables_count": len(extracted.get("tables", [])),
        "images_count": len(extracted.get("images", [])),
        "uploaded_by": username  # Keep this
    }
    
    # ... rest of the function remains the same ...
    
    # Save full text
    joined_text = "\n\n".join([p["text"] for p in extracted.get("texts", [])])
    text_path = os.path.join(doc_dirs["texts"], "full_text.txt")
    with open(text_path, "w", encoding="utf-8") as f:
        f.write(joined_text)
    doc_meta["text_path"] = text_path
    
    # Create chunks
    chunks = []
    
    # Text chunks
    for t in extracted.get("texts", []):
        page_no = t.get("page", 1)
        text = t.get("text", "")
        text_chunks = chunk_text_words(text)
        for i, ch in enumerate(text_chunks):
            chunks.append({
                "chunk_id": str(uuid.uuid4()),
                "doc_id": doc_id,
                "filename": filename,
                "type": "text",
                "page": page_no,
                "chunk_index": i+1,
                "content": ch,
            })
    
    # Table chunks
    for table_obj in extracted.get("tables", []):
        page_no = table_obj.get("page", table_obj.get("sheet", 1))
        csv_path = table_obj.get("csv_path")
        df = table_obj.get("df")
        if df is not None:
            table_text = f"Table ({os.path.basename(csv_path)}):\n" + df.head(20).to_string(index=False)
        else:
            table_text = str(table_obj.get("data", ""))
        chunks.append({
            "chunk_id": str(uuid.uuid4()),
            "doc_id": doc_id,
            "filename": filename,
            "type": "table",
            "page": page_no,
            "content": table_text,
            "csv_path": csv_path
        })
    

    for img in extracted.get("images", []):
        page_no = img.get("page", 1)
        path = img.get("path")
        caption = generate_caption(path) if USE_IMAGE_CAPTIONING else "Image"
        content_text = f"[IMAGE_PATH:{path}]\nCAPTION: {caption}"
        chunks.append({
            "chunk_id": str(uuid.uuid4()),
            "doc_id": doc_id,
            "filename": filename,
            "type": "image",
            "page": page_no,
            "content": content_text,
            "image_path": path,
            "caption": caption
        })
    
    # Save doc metadata
    doc_meta_path = os.path.join(doc_dirs["base"], "doc_meta.json")
    with open(doc_meta_path, "w", encoding="utf-8") as f:
        json.dump(doc_meta, f, indent=2)
    
    return doc_meta, chunks

def save_index_and_chunks(docs_index, all_chunks):
    """Save documents index and all chunks to JSON files."""
    with open(DOCS_INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(docs_index, f, indent=2)
    with open(RAG_CHUNKS_PATH, "w", encoding="utf-8") as f:
        json.dump(all_chunks, f, indent=2)

